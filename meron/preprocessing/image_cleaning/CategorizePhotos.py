import os
import cv2
from openpyxl import Workbook
from shutil import copyfile, rmtree

#based on different HAAR classifiers, categorize photos

photo_folder = "C:\Users\Hussein Lightwalla\Desktop\meron_pics\{}".format('tana_river_meron_survey')
#photo_folder = "C:\Users\Hussein Lightwalla\Desktop\estcascadepic"
rejected_photos_folder = 'C:\Users\Hussein Lightwalla\Desktop\meron_rejected_pics'
approved_photos_folder = 'C:\Users\Hussein Lightwalla\Desktop\meron_approved_pics'
output_folder = 'C:\Users\Hussein Lightwalla\Desktop\meron_combined'
county = 'TanaRiver'

cascPath_eye = 'C:\Users\Hussein Lightwalla\Desktop\cascades\haarcascade_eye_new.xml'
cascPath_face = 'C:\Users\Hussein Lightwalla\Desktop\cascades\haarcascade_frontalface_default.xml'
faceCascade = cv2.CascadeClassifier(cascPath_face)
eyeCascade = cv2.CascadeClassifier(cascPath_eye)

wb = Workbook()
# grab the active worksheet
ws = wb.active

ws['A1'] = "Photo"
ws['B1'] = "Eyes number"
ws['C1'] = "Usable in MERON?"
ws['D1'] = "Reason"
ws['E1'] = "Y differential"

counter = 1
counter_rejected = 0
counter_approved = 0
possible_approved = False
for subdir, dirs, files in os.walk(photo_folder):
    for file in files:
        try:
            analysis_result = ''
            possible_approved = False
            counter = counter + 1
            ws['A{}'.format(counter)] = file
            imagePath = os.path.join(subdir, file)
            image = cv2.imread(imagePath)
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

            faces = faceCascade.detectMultiScale(gray, 1.3, 5)
            if len(faces) == 1:
                counter_approved= counter_approved + 1
                analysis_result = 'Favourable result using Frontal Face classifier (one face detected)'
                copyfile(imagePath, '{}\{}'.format(approved_photos_folder, file))
                ws['C{}'.format(counter)] = 'Yes'
                print 'approved {}'.format(file)
                continue
            else:
                eyes = eyeCascade.detectMultiScale(gray, 1.3, 5)
                if len(eyes) != 2:
                    eyes = eyeCascade.detectMultiScale(
                        gray,
                        scaleFactor=1.115,
                        minNeighbors=1,
                        minSize=(50, 50),
                        maxSize=(250, 250),
                        flags=cv2.CASCADE_SCALE_IMAGE
                    )
                ws['B{}'.format(counter)] = len(eyes)
                if len(eyes) == 2:
                    analysis_result = 'Perfect match with eye classifier'
                if len(eyes) > 2:
                    #check for Y axis differential
                    count_diff = 0
                    diff_y = -999
                    for (x, y, w, h) in eyes:
                        count_diff = count_diff + 1
                        if diff_y == -999:
                            diff_y = y
                        else:
                            diff_y = diff_y - y
                        if diff_y < 0:
                            diff_y = diff_y * -1
                        if count_diff == 2:
                            if diff_y >= 0 and diff_y <=80 :
                                possible_approved = True
                                analysis_result = 'Highly possible to be acceptable photo due to favourable Y axis differential of {}'.format(diff_y)
                                print analysis_result
                            break
                    eyes = faceCascade.detectMultiScale(
                        gray,
                        scaleFactor=1.115,
                        minNeighbors=5,
                        minSize=(50, 50),
                        maxSize=(250, 250),
                        flags=cv2.CASCADE_SCALE_IMAGE
                    )
                if len(eyes) == 2:
                    analysis_result = 'Favourable result found after setting minNeighbours to 5 from 1'
                if len(eyes) != 2:
                    eyes = faceCascade.detectMultiScale(
                        gray,
                        scaleFactor=1.115,
                        minNeighbors=10,
                        minSize=(50, 50),
                        maxSize=(250, 250),
                        flags=cv2.CASCADE_SCALE_IMAGE
                    )
                if len(eyes) == 2:
                    analysis_result = 'Favourable result found after setting minNeighbours to 10 from 5'
                if len(eyes) != 2:
                    eyes = faceCascade.detectMultiScale(
                        gray,
                        scaleFactor=1.115,
                        minNeighbors=30,
                        minSize=(50, 50),
                        maxSize=(250, 250),
                        flags=cv2.CASCADE_SCALE_IMAGE
                    )
                if len(eyes) == 2:
                    analysis_result = 'Favourable result found after setting minNeighbours to 30 from 10'
                if len(eyes) != 2:
                    eyes = faceCascade.detectMultiScale(
                        gray,
                        scaleFactor=1.115,
                        minNeighbors=60,
                        minSize=(50, 50),
                        maxSize=(250, 250),
                        flags=cv2.CASCADE_SCALE_IMAGE
                    )
                if len(eyes) == 2:
                    analysis_result = 'Favourable result found after setting minNeighbours to 60 from 30'
                if len(eyes) == 2 or possible_approved == True:
                    diff_y = -999
                    for (x, y, w, h) in eyes:
                        if diff_y == -999:
                            diff_y = y
                        else:
                            diff_y = diff_y - y
                    counter_approved = counter_approved + 1
                    copyfile(imagePath, '{}\{}'.format(approved_photos_folder, file))
                    ws['C{}'.format(counter)] = 'Yes'
                    if diff_y < 0:
                        diff_y = diff_y * -1
                    ws['E{}'.format(counter)] = diff_y
                    print 'approved {}'.format(file)
                else:
                    counter_rejected = counter_rejected + 1
                    ws['C{}'.format(counter)] = 'No'
                    copyfile(imagePath, '{}\{}'.format(rejected_photos_folder, file))
                    if possible_approved == False:
                        analysis_result = 'HAAR Eye classifier detected more or less than one eye in photo'
                    print 'rejected {}'.format(file)
        except Exception:
            print 'error'
        ws['D{}'.format(counter)] = analysis_result

wb.save('{}\{}'.format(output_folder, '{}_photoanalysis.xlsx'.format(county)))

print('{} approved and {} rejected pics'.format(counter_approved, counter_rejected))

# Draw a rectangle around the faces
#for (x, y, w, h) in faces:
    #cv2.rectangle(image, (x, y), (x+w, y+h), (0, 255, 0), 2)

#cv2.imshow("Faces found", image)
#cv2.waitKey(0)