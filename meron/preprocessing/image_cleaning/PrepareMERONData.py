#for the input excel sheet (SMART data), extract anthros, extract the MERON picture name and output to a new Excel file
from __future__ import division
from openpyxl import Workbook
import pandas as pd
import requests
from datetime import datetime

turkana_meron_survey = 537
marsabit_meron_survey = 513
counter_problematic = 0

form_id = turkana_meron_survey

output_folder = 'C:\Users\Hussein Lightwalla\Desktop\meron_combined'

smart_input_folder = 'C:\Users\Hussein Lightwalla\Desktop\smart_data'

#Marsabit
county = 'Marsabit'
smart_input_identifier_file = 'marsabit_identifier.csv'
smart_input_roster_file = 'marsabit_smart_survey_members.csv'
meron_village_sub_county_area = 'section1/village_sub_county_area'
meron_village = 'section1/village'
meron_village_area = 'section1/village'
meron_cluster = 'section1/cluster'
meron_household = 'section1/household'
meron_team_no = 'section1/team_no'
meron_section_roster = 'section2/repeat_section2'
meron_child_name = 'child_name'
meron_child_picture = 'child_picture'
meron_start_time_photo = 'time_start_photo_capture'
meron_end_time_photo = 'time_end_photo_capture'


#Turkana
if (form_id == turkana_meron_survey):
    county = 'Turkana'
    smart_input_identifier_file = 'turkana_identifier_final.csv'
    smart_input_roster_file = 'turkana_smart_survey_members_final.csv'
    meron_village_sub_county_area = 'smart_identification1/village_area_Sub_County_Area'
    meron_village = 'smart_identification1/village_name'
    meron_village_area = 'smart_identification1/village_area'
    meron_cluster = 'smart_identification_id/cluster_no'
    meron_household = 'smart_identification_id/hh_no'
    meron_team_no = 'smart_identification_id/team_no'
    meron_section_roster = 'household_roster/repeat_section2'
    meron_child_name = 'child_name'
    meron_child_picture = 'child_picture'
    meron_start_time_photo = 'time_start_photo_capture'
    meron_end_time_photo = 'time_end_photo_capture'

df_smart_data_identifier = pd.read_csv('{}\{}'.format(smart_input_folder, smart_input_identifier_file))
df_smart_data_roster = pd.read_csv('{}\{}'.format(smart_input_folder, smart_input_roster_file))

merged_smart_data = pd.merge(left=df_smart_data_identifier,right=df_smart_data_roster, left_on='KEY', right_on='PARENT_KEY')
merged_smart_data.rename(columns={'smart_identification1-village_name':'village_name'}, inplace=True)
merged_smart_data.rename(columns={'smart_identification1-Village_Area':'village_area'}, inplace=True)
merged_smart_data.rename(columns={'smart_identification1-village_area':'village_area'}, inplace=True)
merged_smart_data.rename(columns={'smart_identification_id-cluster_no':'cluster_no'}, inplace=True)
merged_smart_data.rename(columns={'smart_identification_id-hh_no':'hh_no'}, inplace=True)
merged_smart_data.rename(columns={'smart_identification_id-team_no':'team_no'}, inplace=True)
merged_smart_data.village_name = merged_smart_data.village_name.str.lower()
merged_smart_data.q2_2 = merged_smart_data.q2_2.str.lower()


wb = Workbook()
# grab the active worksheet
ws = wb.active
# Data can be assigned directly to cells
ws['A1'] = "County"
ws['B1'] = "Village MERON"
ws['C1'] = "Cluster"
ws['D1'] = "Household ID"
ws['E1'] = "Team ID"
ws['F1'] = "Child ID"
ws['G1'] = "Photo folder ID"
ws['H1'] = "Photo ID"
ws['I1'] = "Child name - MERON"
ws['J1'] = "Child name - SMART"
ws['K1'] = "Weight(KG)"
ws['L1'] = "Length or Height (CM)"
ws['M1'] = "MUAC in CM"
ws['N1'] = "Number of children for this HH in SMART"
ws['O1'] = "Number of children for this HH in MERON"
ws['P1'] = "MERON Data ID"
ws['Q1'] = "Start time for photo capture"
ws['R1'] = "End time for photo capture"
ws['S1'] = "Duration in minutes"
ws['T1'] = "Number of image attachments in MERON submission"
ws['U1'] = "Notes"
ws['V1'] = "Start time of questionnaire SMART"
ws['W1'] = "End time of questionnaire SMART"
ws['X1'] = "Diff between start times of SMART and MERON photo capture"




r = requests.get('https://kobo.kimetrica.com/kobocat/api/v1/data/{}'.format(form_id), auth=('meron_surveys', 'M3ron2018'))
r_json = r.json()
counter = 1
submission_counter = 0
list_households_done = []
for submission in r_json:
    smart_child_start_index = 0
    notes = "Exact match"
    submission_counter = submission_counter + 1
    if submission['_uuid'] == 'd9d33fdc-c0a1-4398-aec7-a9af93f8a77b':
        r = "r"
        #break
    member_no = 0
    if len(submission['_attachments']) > 0:
        member_no = 0
        sub_county = submission[meron_village_sub_county_area]
        village_name = submission[meron_village]
        village_area = submission[meron_village_area]
        cluster_id = submission[meron_cluster]
        hh_id = submission[meron_household]
        team_id = submission[meron_team_no]
        uuid = submission['_uuid']
        key = '{}{}{}{}'.format(sub_county, cluster_id, hh_id, team_id)
        if key in list_households_done:
            smart_child_start_index = list_households_done.count(key)
        list_households_done.append(key)
        print submission_counter
        query = 'village_name == "{}" and cluster_no == "{}" and hh_no == "{}" and team_no == "{}" and q2_1 == "1" and q2_3_grp1_age >= 6 and q2_3_grp1_age <= 59.99 and member_present == "1"'.format((village_name).lower(), cluster_id, hh_id, team_id)
        member_info = merged_smart_data.query(query)
        if member_info.empty == True:
            query = 'village_area == "{}" and cluster_no == "{}" and hh_no == "{}" and team_no == "{}" and q2_1 == "1" and q2_3_grp1_age >= 6 and q2_3_grp1_age <= 59.99 and member_present == "1"'.format(
                (village_area).lower(), cluster_id, hh_id, team_id)
            member_info = merged_smart_data.query(query)
            if member_info.empty == False:
                notes = 'Child name was found but the village name in MERON was village area in SMART : village_name = {}'.format(member_info.iloc[member_no]['village_area'])
        if member_info.empty == True and "_" in village_name:
            village_name = village_name.replace("_", " ")
            query = 'village_name == "{}" and cluster_no == "{}" and hh_no == "{}" and team_no == "{}" and q2_1 == "1" and q2_3_grp1_age >= 6 and q2_3_grp1_age <= 59.99 and member_present == "1"'.format(
                (village_name).lower(), cluster_id, hh_id, team_id)
            member_info = merged_smart_data.query(query)
        if member_info.empty == True:
            #Maybe they put the wrong team number
            query = 'village_area == "{}" and cluster_no == "{}" and hh_no == "{}" and q2_1 == "1" and q2_3_grp1_age >= 6 and q2_3_grp1_age <= 59.99 and member_present == "1"'.format(
                (village_name).lower(), cluster_id, hh_id)
            member_info = merged_smart_data.query(query)
            if member_info.empty == False:
                notes = 'Child name was found but the team number seems to have been entered wrongly in either MERON or SMART'
        if member_info.empty == True:
            child_name = submission[meron_section_roster][0]['{}/{}'.format(meron_section_roster, meron_child_name)]
            child_name = child_name.replace('\n', ' ').replace('\r', '').replace('\t', '').encode('utf-8').strip()
            query = 'cluster_no == "{}" and hh_no == "{}" and team_no == "{}" and q2_1 == "1" and q2_3_grp1_age >= 6 and q2_3_grp1_age <= 59.99 and q2_2 == "{}" and member_present == "1"'.format(cluster_id, hh_id, team_id, child_name.lower())
            print query
            member_info = merged_smart_data.query(query)
            if member_info.empty == False:
                notes = 'Child name was found but associated with different village : village_name = {}'.format(member_info.iloc[member_no]['village_name'])
        if member_info.empty == True:
            child_name = submission[meron_section_roster][0]['{}/{}'.format(meron_section_roster, meron_child_name)]
            child_name = child_name.replace('\n', ' ').replace('\r', '').replace('\t', '').encode('utf-8').strip()
            query = 'q2_2 == "{}"'.format(child_name.lower())
            member_info = merged_smart_data.query(query)
            if member_info.empty == False:
                notes = 'Child name was found but associated with different identifiers : village_name = {} cluster_no = {} hh_no = {} team_no = {}'.format\
                    (member_info.iloc[member_no]['village_name'],member_info.iloc[member_no]['cluster_no'],member_info.iloc[member_no]['hh_no'],member_info.iloc[member_no]['team_no'])
        count_children_SMART = len(member_info.index)
        count_children_MERON = len(submission[meron_section_roster])
        if count_children_MERON == count_children_SMART:
            smart_child_start_index = 0
        loop_counter = count_children_MERON
        #loop_counter = count_children_SMART
        #if count_children_MERON > count_children_SMART:
        #    loop_counter = count_children_MERON
        if member_info.empty == False or count_children_MERON > 0:
            member_info = member_info.sort_values(by=['member_number'])
            for x in range(0, loop_counter):
                counter = counter + 1
                ws['A{}'.format(counter)] = county
                ws['B{}'.format(counter)] = village_name
                ws['C{}'.format(counter)] = cluster_id
                ws['D{}'.format(counter)] = hh_id
                ws['E{}'.format(counter)] = team_id
                ws['F{}'.format(counter)] = member_no + 1
                ws['G{}'.format(counter)] = uuid
                ws['T{}'.format(counter)] = len(submission['_attachments'])
                SMART_child_name = ''
                MERON_child_name = ''
                time_start_SMART = None
                time_start_MERON = None
                try:
                    ws['H{}'.format(counter)] = submission[meron_section_roster][member_no]['{}/{}'.format(meron_section_roster, meron_child_picture)]
                    MERON_child_name = submission[meron_section_roster][member_no]['{}/{}'.format(meron_section_roster, meron_child_name)].replace('\n', ' ').replace('\r', '').replace('\t', '')
                    ws['I{}'.format(counter)] = MERON_child_name
                    ws['R{}'.format(counter)] = submission[meron_section_roster][member_no]['{}/{}'.format(meron_section_roster, meron_end_time_photo)]
                    start_time = submission[meron_section_roster][member_no]['{}/{}'.format(meron_section_roster, meron_end_time_photo)][:8]
                    end_time = submission[meron_section_roster][member_no]['{}/{}'.format(meron_section_roster, meron_end_time_photo)][:8]
                    if '{}/{}'.format(meron_section_roster, meron_start_time_photo) in submission[meron_section_roster][member_no]:
                        ws['Q{}'.format(counter)] = submission[meron_section_roster][member_no]['{}/{}'.format(meron_section_roster, meron_start_time_photo)]
                        start_time = submission[meron_section_roster][member_no]['{}/{}'.format(meron_section_roster, meron_start_time_photo)][:8]
                        time_start_MERON = start_time
                    FMT = '%H:%M:%S'
                    tdelta = datetime.strptime(end_time, FMT) - datetime.strptime(start_time, FMT)
                    ws['S{}'.format(counter)] = tdelta.seconds / 60
                except IndexError:
                    notes = "No matching record in MERON for this child"
                    counter_problematic = counter_problematic + 1
                    print "MERON index error with " + uuid
                try:
                    if (member_info.iloc[member_no + smart_child_start_index].empty == False):
                        village_name_SMART = member_info.iloc[member_no + smart_child_start_index]['village_name']
                        SMART_child_name = member_info.iloc[member_no + smart_child_start_index]['q2_2'].replace('\n', ' ').replace('\r', '').replace('\t', '')
                        ws['J{}'.format(counter)] = SMART_child_name
                        ws['K{}'.format(counter)] = member_info.iloc[member_no + smart_child_start_index]['member_present_group-child_health_n_nutrition-child_6_59-child_anthropometry_1-q3_1_f']
                        ws['L{}'.format(counter)] = member_info.iloc[member_no + smart_child_start_index]['member_present_group-child_health_n_nutrition-child_6_59-child_anthropometry_1-q3_1_g']
                        ws['M{}'.format(counter)] = member_info.iloc[member_no + smart_child_start_index]['member_present_group-child_health_n_nutrition-child_6_59-child_anthropometry_1-q3_1_i']
                        ws['V{}'.format(counter)] = member_info.iloc[member_no + smart_child_start_index]['start']
                        time_start_SMART = ws['V{}'.format(counter)] = member_info.iloc[member_no + smart_child_start_index]['start'][12:]
                        ws['W{}'.format(counter)] = member_info.iloc[member_no + smart_child_start_index]['end']
                except IndexError:
                    notes = "No matching record in SMART for this child"
                    counter_problematic = counter_problematic + 1
                    print "SMART index error with " + uuid
                ws['N{}'.format(counter)] = count_children_SMART
                ws['O{}'.format(counter)] = count_children_MERON
                ws['P{}'.format(counter)] = submission['_id']
                ws['U{}'.format(counter)] = notes
                if time_start_MERON is not None and time_start_SMART is not None:
                    FMT = '%H:%M:%S'
                    tdelta = datetime.strptime(time_start_SMART, FMT) - datetime.strptime(time_start_MERON, FMT)
                    ws['X{}'.format(counter)] = tdelta.seconds
                member_no = member_no + 1
# Save the file
wb.save('{}\{}'.format(output_folder,'{}.xlsx'.format(county)))

print '{} submissions have mismatch issues. {}% of the total.'.format(counter_problematic, ((counter_problematic/submission_counter) * 100))
print '{} submissions are usable for MERON modeling'.format(submission_counter - counter_problematic)

#wb = openpyxl.load_workbook('{}\{}'.format(folder,file_backcheck))
#sheet = wb['Sheet1']
#print sheet['A1'].value
